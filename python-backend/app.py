import os
import re
import warnings
from flask import Flask, request, jsonify
from flask_cors import CORS  # type: ignore[import-untyped]

warnings.filterwarnings(
    "ignore",
    message="All support for the `google.generativeai` package has ended",
    category=FutureWarning,
)
import google.generativeai as genai
from werkzeug.utils import secure_filename
import uuid
import json
from datetime import datetime
import mimetypes
from PIL import Image
import PyPDF2
from docx import Document
import openpyxl  # type: ignore[import-untyped]
from dotenv import load_dotenv
from typing import Any, Optional, Protocol, cast

# Load environment variables
load_dotenv()

# RAG stack (sentence-transformers / huggingface) can fail on some Python or dependency combos — app still runs without it.
RAGSystem: Any = None
try:
    from rag_system import RAGSystem as _RAGSystem  # type: ignore[import-not-found]

    RAGSystem = _RAGSystem
except Exception as rag_import_err:
    print(f"⚠️  RAG module not loaded (upload + Gemini + local Q&A still work): {rag_import_err}")

app = Flask(__name__)
CORS(app)  # type: ignore[arg-type]

# Configure Gemini API (never commit real keys — use python-backend/.env)
GEMINI_API_KEY = (os.getenv("GEMINI_API_KEY") or "").strip()
model: Any = None
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)  # pyright: ignore[reportPrivateImportUsage]
    model = genai.GenerativeModel("gemini-2.0-flash")  # pyright: ignore[reportPrivateImportUsage]

class RAGSystemProtocol(Protocol):
    embedding_dimension: int

    def add_document(self, document_id: str, content: str, metadata: dict[str, Any]) -> None: ...
    def ask_question(self, question: str, document_id: str) -> dict[str, Any]: ...
    def get_document_count(self) -> int: ...
    def get_total_chunks(self) -> int: ...
    def build_vector_index(self, payload: Any) -> Any: ...
    def generate_rag_response(
        self, question: str, filename: str, document_id: str | None = None
    ) -> Any: ...

    def remove_document(self, document_id: str) -> bool: ...


rag_system: Optional[RAGSystemProtocol] = None
if GEMINI_API_KEY and RAGSystem is not None:
    try:
        rag_system = RAGSystem(api_key=GEMINI_API_KEY)
        print("✅ RAG System initialized successfully")
    except Exception as e:
        print(f"⚠️  RAG system init failed: {e}")
        rag_system = None

# Configuration
UPLOAD_FOLDER = 'uploads'
USERS_DATA_FOLDER = 'users_data'
ALLOWED_EXTENSIONS = {
    'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 
    'xls', 'xlsx', 'csv', 'ppt', 'pptx', 'bmp', 'tiff', 'svg'
}

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(USERS_DATA_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_user_data_path(user_id):
    """Get the path to user's data directory"""
    user_dir = os.path.join(USERS_DATA_FOLDER, user_id)
    os.makedirs(user_dir, exist_ok=True)
    return user_dir

def get_predefined_response(question_type, question):
    """Get predefined responses for common questions"""
    import random
    
    responses = {
        'greeting': [
            "Hello! I'm here to help you analyze your documents. What would you like to know about your uploaded files?",
            "Hi there! I can help you understand and explore the content of your documents. How can I assist you today?",
            "Good day! I'm your document analysis assistant. Feel free to ask me questions about your uploaded files.",
            "Hello! Welcome to your document Q&A system. I'm ready to help you explore your content."
        ],
        'conversational': {
            'who are you': "I'm your personal document analysis assistant, powered by Gemini AI. I help you understand and explore the content of your uploaded documents.",
            'what can you do': "I can analyze your uploaded documents, answer questions about their content, extract specific information, and help you understand complex data. Just ask me anything about your files!",
            'how are you': "I'm doing great and ready to help you explore your documents! What would you like to know about your files?",
            'what is your name': "I'm your document analysis assistant! I don't have a specific name, but I'm here to help you with all your document questions.",
            'can you help': "Absolutely! I'm here to help you understand and analyze your documents. What would you like to know?",
            'help me': "Of course! I'm ready to help you explore your documents. What specific information are you looking for?"
        },
        'polite': {
            'thank you': "You're very welcome! I'm happy to help you with your documents. Is there anything else you'd like to know?",
            'thanks': "My pleasure! Feel free to ask more questions about your documents anytime.",
            'please': "Of course! I'm here to help you understand your documents. What would you like to know?",
            'excuse me': "No problem at all! How can I help you with your documents today?"
        },
        'goodbye': [
            "Goodbye! Feel free to come back anytime you need help with your documents.",
            "Thanks for using the document analysis system! Have a great day!",
            "It was my pleasure helping you explore your documents. See you next time!",
            "Farewell! Your documents will be here whenever you need to analyze them again."
        ]
    }
    
    question_lower = question.lower().strip()
    
    if question_type == 'greeting':
        return random.choice(responses['greeting'])
    elif question_type == 'conversational':
        for key, response in responses['conversational'].items():
            if key in question_lower:
                return response
        return "I'm your document analysis assistant! I can help you understand and explore the content of your uploaded files."
    elif question_type == 'polite':
        for key, response in responses['polite'].items():
            if key in question_lower:
                return response
        return "Thank you! I'm happy to help you with your documents. What would you like to know?"
    elif question_type == 'goodbye':
        return random.choice(responses['goodbye'])
    
    return None

def categorize_question(question):
    """Categorize the type of question being asked"""
    question_lower = question.lower().strip()
    
    # Check for document-related keywords first (highest priority)
    document_keywords = [
        'document', 'file', 'content', 'text', 'information', 'data', 'report', 'analysis',
        'what is', 'what are', 'tell me about', 'explain', 'describe', 'summary', 
        'technologies', 'project', 'details', 'findings', 'results', 'conclusion'
    ]
    
    # Check if it's clearly a document question
    if any(keyword in question_lower for keyword in document_keywords):
        # But exclude meta questions about the AI itself
        meta_phrases = ['who are you', 'what are you exactly', 'what is your name', 'introduce yourself']
        if not any(meta in question_lower for meta in meta_phrases):
            return 'document_question'
    
    # Simple greetings (exact matches)
    simple_greetings = ['hello', 'hi', 'hey', 'good morning', 'good afternoon', 'good evening']
    if question_lower in simple_greetings or question_lower.startswith(tuple(simple_greetings)):
        return 'greeting'
    
    # Polite expressions
    if question_lower in ['thanks', 'thank you', 'please', 'excuse me']:
        return 'polite'
    
    # Meta/conversational questions about the AI
    meta_questions = ['who are you', 'what can you do', 'what is your name', 'introduce yourself', 'how are you']
    if any(meta in question_lower for meta in meta_questions):
        return 'conversational'
    
    # Goodbye expressions
    goodbyes = ['bye', 'goodbye', 'see you', 'farewell']
    if any(bye in question_lower for bye in goodbyes):
        return 'goodbye'
    
    # Default to document question
    return 'document_question'

def extract_text_from_file(file_path, file_type):
    """Extract text from various file types"""
    try:
        if file_type.startswith('image/'):
            # For images, we'll use Gemini API for OCR
            return None  # Will be processed by Gemini
        
        elif file_type == 'application/pdf':
            print("Extracting text from PDF...")
            try:
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    text = ""
                    for page_num in range(len(pdf_reader.pages)):
                        page = pdf_reader.pages[page_num]
                        page_text = page.extract_text()
                        text += page_text + "\n"
                    print(f"PDF text extraction successful. Length: {len(text)} characters")
                    return text if text.strip() else "PDF processed but no readable text found"
            except Exception as pdf_error:
                print(f"PDF extraction error: {str(pdf_error)}")
                return f"PDF processing error: {str(pdf_error)}"
        
        elif file_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/msword']:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        
        elif file_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
            return text
        
        elif file_type.startswith('text/'):
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        
        else:
            return "Unsupported file type for text extraction"
    
    except Exception as e:
        return f"Error extracting text: {str(e)}"

def process_with_gemini(file_path, file_type, extracted_text=None):
    """Process file using Gemini API for comprehensive analysis"""
    print(f"Processing file: {file_path}, type: {file_type}")

    if model is None:
        print("Gemini skipped: set GEMINI_API_KEY in python-backend/.env")
        return {
            "error": "Gemini API key not set. Add GEMINI_API_KEY to python-backend/.env and restart the server.",
            "extracted_text": extracted_text if extracted_text else None,
            "file_type": file_type,
            "processing_method": "skipped_no_api_key",
        }

    try:
        if file_type.startswith('image/'):
            print("Processing as image with Gemini Vision...")
            # For images, use Gemini's vision capabilities
            image = Image.open(file_path)
            
            prompt = """
            Analyze this image and extract all information:
            1. Any text present (OCR)
            2. Visual description
            3. Objects or elements identified
            4. Context and purpose
            
            Provide your analysis in clear text format.
            """
            
            response = model.generate_content([prompt, image])
            print("Gemini vision processing successful")
            return {
                "ocr_text": response.text,
                "file_type": "image",
                "processing_method": "gemini_vision"
            }
        
        else:
            print("Processing as document with text analysis...")
            # For other files, use the extracted text with Gemini for analysis
            if extracted_text and len(extracted_text.strip()) > 0:
                # Limit text to prevent token issues
                text_sample = extracted_text[:8000] if len(extracted_text) > 8000 else extracted_text
                
                prompt = f"""
                Analyze this document content:
                
                {text_sample}
                
                Provide:
                1. Document summary
                2. Key topics
                3. Important information
                4. Main conclusions
                
                Keep response concise and clear.
                """
                
                response = model.generate_content(prompt)
                print("Gemini text analysis successful")
                return {
                    "extracted_text": extracted_text,
                    "gemini_analysis": response.text,
                    "file_type": "document",
                    "processing_method": "gemini_text_analysis"
                }
            else:
                print("No text content available for analysis")
                return {
                    "message": "File processed but no text content found",
                    "file_type": file_type,
                    "processing_method": "no_text_content"
                }
    
    except Exception as e:
        print(f"Gemini processing error: {str(e)}")
        return {
            "error": f"Gemini processing failed: {str(e)}",
            "extracted_text": extracted_text if extracted_text else None,
            "file_type": file_type,
            "processing_method": "error"
        }

def save_document_data(user_id, document_data):
    """Save document data to user-specific JSON file"""
    user_dir = get_user_data_path(user_id)
    user_library_file = os.path.join(user_dir, 'library.json')
    
    # Load existing library or create new
    if os.path.exists(user_library_file):
        with open(user_library_file, 'r', encoding='utf-8') as f:
            library = json.load(f)
    else:
        library = {
            "user_id": user_id,
            "created_at": datetime.now().isoformat(),
            "documents": [],
            "total_documents": 0
        }
    
    # Add new document
    documents = cast(list[Any], library["documents"])
    documents.append(document_data)
    library["total_documents"] = len(documents)
    library["last_updated"] = datetime.now().isoformat()
    
    # Save updated library
    with open(user_library_file, 'w', encoding='utf-8') as f:
        json.dump(library, f, indent=2, ensure_ascii=False)
    
    # Also save individual document file
    doc_file = os.path.join(user_dir, f"{document_data['document_id']}.json")
    with open(doc_file, 'w', encoding='utf-8') as f:
        json.dump(document_data, f, indent=2, ensure_ascii=False)
    
    return library

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "healthy",
        "message": "RAG Backend API is running",
        "timestamp": datetime.now().isoformat()
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Upload and process file endpoint"""
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        user_id = request.form.get('user_id')
        
        if not user_id:
            return jsonify({"error": "User ID is required"}), 400
        
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        if file and allowed_file(file.filename):
            # Generate unique filename
            file_id = str(uuid.uuid4())
            raw_filename = file.filename or ""
            filename = secure_filename(raw_filename)
            file_extension = filename.rsplit('.', 1)[1].lower()
            unique_filename = f"{file_id}.{file_extension}"
            
            # Save file
            file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
            file.save(file_path)
            
            # Get file type
            file_type = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
            file_size = os.path.getsize(file_path)
            
            # Extract text (if applicable)
            print(f"Extracting text from {filename}...")
            extracted_text = extract_text_from_file(file_path, file_type)
            print(f"Text extraction result: {len(extracted_text) if extracted_text else 0} characters")
            
            # Process with Gemini API
            print("Starting Gemini API processing...")
            gemini_result = process_with_gemini(file_path, file_type, extracted_text)
            print(f"Gemini processing complete: {gemini_result.get('processing_method', 'unknown')}")
            
            # Process with RAG system for better Q&A capabilities
            print("Processing document with RAG system...")
            rag_processed = False
            if rag_system and extracted_text and extracted_text.strip():
                try:
                    # Build FAISS vector index for this document
                    rag_result = rag_system.build_vector_index({
                        "document_id": file_id,
                        "original_filename": filename,
                        "file_type": file_type,
                        "upload_timestamp": datetime.now().isoformat(),
                        "extracted_text": extracted_text,
                        "gemini_processing": gemini_result
                    })
                    if rag_result.get("success"):
                        rag_processed = True
                        print(f"✅ Document indexed in RAG: {rag_result.get('chunks_created')} chunks")
                    else:
                        print(f"RAG indexing issue: {rag_result.get('error')}")
                except Exception as e:
                    print(f"RAG processing failed: {str(e)}")
            
            # Prepare document data
            document_data = {
                "document_id": file_id,
                "original_filename": filename,
                "stored_filename": unique_filename,
                "file_path": file_path,
                "file_type": file_type,
                "file_size": file_size,
                "upload_timestamp": datetime.now().isoformat(),
                "user_id": user_id,
                "extracted_text": extracted_text,
                "gemini_processing": gemini_result,
                "rag_processed": rag_processed,
                "metadata": {
                    "file_extension": file_extension,
                    "processing_status": (
                        "completed" if "error" not in gemini_result and 
                        ("gemini_analysis" in gemini_result or "ocr_text" in gemini_result)
                        else "error_with_fallback" if "error" in gemini_result
                        else "processed"
                    ),
                    "has_text": bool(extracted_text and extracted_text.strip()),
                    "has_gemini_analysis": "gemini_analysis" in gemini_result or "ocr_text" in gemini_result,
                    "rag_enabled": rag_processed,
                    "error_details": gemini_result.get("error", None) if "error" in gemini_result else None
                }
            }
            
            # Save to user's library
            updated_library = save_document_data(user_id, document_data)
            
            return jsonify({
                "success": True,
                "message": "File uploaded and processed successfully",
                "document_id": file_id,
                "filename": filename,
                "processing_result": gemini_result,
                "library_stats": {
                    "total_documents": updated_library["total_documents"],
                    "last_updated": updated_library["last_updated"]
                }
            })
        
        else:
            return jsonify({"error": "File type not allowed"}), 400
    
    except Exception as e:
        return jsonify({"error": f"Upload failed: {str(e)}"}), 500

@app.route('/api/library/<user_id>', methods=['GET'])
def get_user_library(user_id):
    """Get user's document library"""
    try:
        user_dir = get_user_data_path(user_id)
        library_file = os.path.join(user_dir, 'library.json')
        
        if os.path.exists(library_file):
            with open(library_file, 'r', encoding='utf-8') as f:
                library = json.load(f)
            return jsonify(library)
        else:
            return jsonify({
                "user_id": user_id,
                "documents": [],
                "total_documents": 0,
                "message": "No library found for user"
            })
    
    except Exception as e:
        return jsonify({"error": f"Failed to retrieve library: {str(e)}"}), 500

@app.route('/api/document/<user_id>/<document_id>', methods=['GET'])
def get_document(user_id, document_id):
    """Get specific document details"""
    try:
        user_dir = get_user_data_path(user_id)
        doc_file = os.path.join(user_dir, f"{document_id}.json")
        
        if os.path.exists(doc_file):
            with open(doc_file, 'r', encoding='utf-8') as f:
                document = json.load(f)
            return jsonify(document)
        else:
            return jsonify({"error": "Document not found"}), 404
    
    except Exception as e:
        return jsonify({"error": f"Failed to retrieve document: {str(e)}"}), 500

@app.route('/api/search/<user_id>', methods=['POST'])
def search_documents(user_id):
    """Search through user's documents"""
    try:
        payload = request.get_json(silent=True) or {}
        search_query = str(payload.get('query', '')).lower()
        
        if not search_query:
            return jsonify({"error": "Search query is required"}), 400
        
        user_dir = get_user_data_path(user_id)
        library_file = os.path.join(user_dir, 'library.json')
        
        if not os.path.exists(library_file):
            return jsonify({
                "results": [],
                "total_results": 0,
                "query": search_query
            })
        
        with open(library_file, 'r', encoding='utf-8') as f:
            library = json.load(f)
        
        # Search through documents
        results = []
        for doc in library["documents"]:
            # Search in filename, extracted text, and Gemini analysis
            searchable_content = [
                doc.get("original_filename", "").lower(),
                doc.get("extracted_text", "").lower(),
                str(doc.get("gemini_processing", {})).lower()
            ]
            
            if any(search_query in content for content in searchable_content):
                results.append(doc)
        
        return jsonify({
            "results": results,
            "total_results": len(results),
            "query": search_query,
            "searched_in": library["total_documents"]
        })
    
    except Exception as e:
        return jsonify({"error": f"Search failed: {str(e)}"}), 500

@app.route('/api/delete/<user_id>/<document_id>', methods=['DELETE'])
def delete_document(user_id, document_id):
    """Delete a document from user's library"""
    try:
        user_dir = get_user_data_path(user_id)
        library_file = os.path.join(user_dir, 'library.json')
        doc_file = os.path.join(user_dir, f"{document_id}.json")
        
        # Load library
        if not os.path.exists(library_file):
            return jsonify({"error": "Library not found"}), 404
        
        with open(library_file, 'r', encoding='utf-8') as f:
            library = json.load(f)
        
        # Find and remove document
        document_found = False
        for i, doc in enumerate(library["documents"]):
            if doc["document_id"] == document_id:
                # Remove uploaded file
                if os.path.exists(doc["file_path"]):
                    os.remove(doc["file_path"])
                
                # Remove from library
                library["documents"].pop(i)
                library["total_documents"] = len(library["documents"])
                library["last_updated"] = datetime.now().isoformat()
                document_found = True
                break
        
        if not document_found:
            return jsonify({"error": "Document not found"}), 404

        if rag_system:
            try:
                rag_system.remove_document(document_id)
            except Exception as rag_del_err:
                print(f"RAG remove_document: {rag_del_err}")

        # Save updated library
        with open(library_file, 'w', encoding='utf-8') as f:
            json.dump(library, f, indent=2, ensure_ascii=False)
        
        # Remove individual document file
        if os.path.exists(doc_file):
            os.remove(doc_file)
        
        return jsonify({
            "success": True,
            "message": "Document deleted successfully",
            "remaining_documents": library["total_documents"]
        })
    
    except Exception as e:
        return jsonify({"error": f"Delete failed: {str(e)}"}), 500

@app.route('/api/qa/get-document-content/<user_id>/<document_id>', methods=['GET'])
def get_document_for_qa(user_id, document_id):
    """Get document content for Q&A"""
    try:
        user_dir = get_user_data_path(user_id)
        doc_file = os.path.join(user_dir, f"{document_id}.json")
        
        if not os.path.exists(doc_file):
            return jsonify({"error": "Document not found"}), 404
        
        with open(doc_file, 'r', encoding='utf-8') as f:
            document = json.load(f)
        
        # Extract content for Q&A
        content = ""
        
        # Add extracted text if available
        if document.get('extracted_text'):
            content += f"Document Text:\n{document['extracted_text']}\n\n"
        
        # Add Gemini analysis
        gemini_processing = document.get('gemini_processing', {})
        if gemini_processing.get('ocr_text'):
            content += f"OCR Content:\n{gemini_processing['ocr_text']}\n\n"
        if gemini_processing.get('gemini_analysis'):
            content += f"Analysis:\n{gemini_processing['gemini_analysis']}\n\n"
        
        return jsonify({
            "document_id": document_id,
            "filename": document.get('original_filename', 'Unknown'),
            "file_type": document.get('file_type', 'unknown'),
            "content": content.strip(),
            "upload_date": document.get('upload_timestamp', ''),
            "has_content": bool(content.strip())
        })
    
    except Exception as e:
        return jsonify({"error": f"Failed to get document content: {str(e)}"}), 500

@app.route('/api/qa/ask', methods=['POST'])
def ask_question():
    """Ask a question about a specific document using RAG system"""
    try:
        data = request.get_json(silent=True) or {}
        user_id = data.get('user_id')
        document_id = data.get('document_id')
        question = data.get('question')
        
        print(f"Q&A Request - User: {user_id}, Doc: {document_id}, Question: {question}")
        
        if not all([user_id, document_id, question]):
            return jsonify({"error": "user_id, document_id, and question are required"}), 400

        user_id = str(user_id)
        document_id = str(document_id)
        question = str(question)
        
        # Get document content
        user_dir = get_user_data_path(user_id)
        doc_file = os.path.join(user_dir, f"{document_id}.json")
        
        if not os.path.exists(doc_file):
            return jsonify({"error": "Document not found"}), 404
        
        with open(doc_file, 'r', encoding='utf-8') as f:
            document = json.load(f)
        
        filename = document.get('original_filename', 'Unknown')
        print(f"Processing Q&A for document: {filename}")
        
        # Categorize the question
        question_type = categorize_question(question)
        print(f"Question type: {question_type}")
        
        # Handle greeting and conversational questions
        rag_metadata = None

        if question_type == 'greeting':
            answer = f"Hello! I'm your advanced document analysis assistant powered by RAG (Retrieval-Augmented Generation). I'm currently ready to help you with questions about '{filename}'. What would you like to know about this document?"
        elif question_type == 'polite':
            answer = f"You're welcome! I'm here to help you analyze '{filename}' using advanced AI techniques. Feel free to ask me anything about this document's content."
        elif question_type == 'conversational':
            answer = f"I'm an AI assistant specialized in document analysis using RAG technology. I can help you understand and extract information from your uploaded documents. Right now, I'm ready to answer questions about '{filename}'. You can ask me about its content, key information, or any specific details you need."
        elif question_type == 'goodbye':
            answer = f"Goodbye! Feel free to come back anytime if you have more questions about '{filename}' or other documents. Have a great day!"
        else:
            # Handle document-related questions with RAG system
            print("Using RAG system for document Q&A...")
            
            # Try to use RAG system if available (even if document wasn't previously processed)
            if rag_system:
                try:
                    # Build vector index for this document (in-memory RAG)
                    build_result = rag_system.build_vector_index(document)
                    if build_result and 'error' in build_result:
                        print(f"RAG index building failed: {build_result['error']}")
                        raise Exception(f"RAG index building failed: {build_result['error']}")
                    
                    # Use RAG system for enhanced Q&A
                    rag_response = rag_system.generate_rag_response(
                        question, filename, document_id
                    )
                    
                    # Check if RAG response has a valid answer
                    answer = rag_response.get('answer', '').strip()
                    generic_no_info = (
                        "I couldn't find relevant information in the document to answer your question. Please try rephrasing your question or ask about content that's actually in the document."
                    )
                    has_error = 'error' in rag_response
                    if not answer or answer == 'No generated text returned for this query.' or answer == generic_no_info:
                        print(f"RAG response was insufficient: {answer}")
                        answer = fallback_document_qa(document, question, filename)
                        rag_metadata = {
                            'rag_enabled': False,
                            'processing_method': 'fallback_document_qa',
                            'rag_error': 'RAG returned no useful answer'
                        }
                    else:
                        rag_info = rag_response.get('rag_info', {})
                        rag_metadata = {
                            'rag_enabled': True,
                            'chunks_retrieved': rag_info.get('chunks_retrieved', 0),
                            'processing_method': 'rag_vector_search',
                            'document_rag_processed': document.get('rag_processed', False),
                            'has_error': has_error
                        }
                        if has_error:
                            rag_metadata['rag_error'] = rag_response['error']
                            print(f"RAG response with extractive fallback: {rag_response['error'][:100]}")
                        else:
                            print(f"RAG response generated with {rag_metadata['chunks_retrieved']} relevant chunks")

                except Exception as rag_error:
                    print(f"RAG system error: {str(rag_error)}")
                    # Fallback to Gemini direct
                    answer = fallback_document_qa(document, question, filename)
                    rag_metadata = {
                        'rag_enabled': False,
                        'processing_method': 'fallback_gemini',
                        'rag_error': str(rag_error)
                    }
            else:
                # Use Gemini directly with full document text
                answer = fallback_document_qa(document, question, filename)
                rag_metadata = {
                    'rag_enabled': False,
                    'processing_method': 'gemini_direct'
                }
        
        # Save Q&A to history
        qa_entry = {
            "question": question,
            "answer": answer,
            "timestamp": datetime.now().isoformat(),
            "document_id": document_id,
            "document_filename": filename,
            "question_type": question_type,
            "rag_metadata": rag_metadata
        }
        
        # Save to user's Q&A history
        qa_history_file = os.path.join(user_dir, 'qa_history.json')
        if os.path.exists(qa_history_file):
            with open(qa_history_file, 'r', encoding='utf-8') as f:
                qa_history = json.load(f)
        else:
            qa_history = {"user_id": user_id, "qa_sessions": []}
        
        qa_sessions = cast(list[Any], qa_history["qa_sessions"])
        qa_sessions.append(qa_entry)
        
        # Keep only last 100 Q&A sessions
        if len(qa_sessions) > 100:
            qa_history["qa_sessions"] = qa_sessions[-100:]
        
        with open(qa_history_file, 'w', encoding='utf-8') as f:
            json.dump(qa_history, f, indent=2, ensure_ascii=False)
        
        print(f"Q&A completed successfully")
        
        return jsonify({
            "question": question,
            "answer": answer,
            "document_filename": filename,
            "timestamp": qa_entry["timestamp"],
            "question_type": question_type,
            "rag_metadata": rag_metadata
        })
    
    except Exception as e:
        print(f"Q&A error: {str(e)}")
        return jsonify({"error": f"Q&A failed: {str(e)}"}), 500

def fallback_document_qa(document, question, filename):
    """Fallback method for Q&A when RAG system fails"""
    # Prepare content for Q&A
    content = ""
    
    # Combine all available content from the JSON
    if document.get('extracted_text'):
        content += f"Extracted Text:\n{document['extracted_text']}\n\n"
    
    gemini_processing = document.get('gemini_processing', {})
    if gemini_processing.get('ocr_text'):
        content += f"OCR Content:\n{gemini_processing['ocr_text']}\n\n"
    if gemini_processing.get('gemini_analysis'):
        content += f"Document Analysis:\n{gemini_processing['gemini_analysis']}\n\n"
    
    # Add metadata information
    metadata = document.get('metadata', {})
    if metadata:
        content += f"File Information:\n"
        content += f"- File Type: {document.get('file_type', 'unknown')}\n"
        content += f"- File Size: {document.get('file_size', 0)} bytes\n"
        content += f"- Upload Date: {document.get('upload_timestamp', 'unknown')}\n\n"
    
    if not content.strip():
        return "I apologize, but there's no readable content available in this document for me to analyze. The document might be corrupted, empty, or in a format that couldn't be processed."

    # Prefer structured extraction from the PDF for ML lifecycle / step lists (accurate, full list).
    body_for_steps = _extract_primary_document_text(content)
    lifecycle_quick = _answer_ml_lifecycle_from_text(
        body_for_steps, question.lower().strip(), filename
    )
    if lifecycle_quick:
        return lifecycle_quick

    # Create comprehensive prompt for document questions with structured formatting
    prompt = f"""
You are a helpful AI assistant that answers questions based ONLY on the provided document content.

CRITICAL INSTRUCTIONS FOR STRUCTURED ANSWERS:
1. Answer concisely and accurately using ONLY the document content below
2. If the information is not in the document, say: "This information is not available in the document"
3. Do not repeat information unnecessarily - be direct and to the point
4. Cite specific details from the document when relevant
5. STRUCTURE YOUR ANSWER CLEARLY:
   - Use bullet points (•) for lists of items, features, or characteristics
   - Use numbered lists (1., 2., 3.) for steps, sequences, or ordered items
   - Use headings (##) for major sections when the answer is complex
   - Use tables when presenting comparative data
   - Bold important terms or key points using **bold** formatting
6. For definitions or formulas, provide them exactly as they appear in the document
7. Avoid introductory phrases like "Based on the document" or "The document says" - just give the answer
8. Always provide a well-organized, structured answer even for simple questions

FORMATTING GUIDELINES:
- Start with a brief direct answer if the question is factual
- For explanatory questions, provide a structured explanation
- For comparison questions, use a table or bullet points
- For step-by-step instructions, use numbered lists
- For lists of items, use bullet points
- Use markdown formatting for better readability

DOCUMENT: {filename}

DOCUMENT CONTENT:
{content[:15000]}  # Limit to prevent token issues

USER QUESTION: {question}

Provide a clear, concise, and well-structured answer with proper formatting:"""

    if model is None:
        return local_document_qa_fallback(content, question, filename)

    try:
        response = model.generate_content(prompt)
        return response.text
    except Exception as gemini_error:
        print(f"Fallback Gemini API error: {str(gemini_error)}")
        # Final non-LLM fallback: return the most relevant document lines so Q&A
        # still works even when Gemini is unavailable (quota/key/network issues).
        return local_document_qa_fallback(content, question, filename)


def _extract_primary_document_text(content: str) -> str:
    """Prefer raw PDF text for structural parsing (numbered lists, headings)."""
    if "Extracted Text:" in content:
        tail = content.split("Extracted Text:", 1)[1]
        for stop in (
            "\n\nOCR Content:",
            "\nOCR Content:",
            "\n\nDocument Analysis:",
            "\nDocument Analysis:",
            "\n\nFile Information:",
            "\nFile Information:",
        ):
            if stop in tail:
                tail = tail.split(stop, 1)[0]
        return tail.strip()
    return content.strip()


def _numbered_steps_from_text(text: str) -> list[str]:
    """Collect 1., 2., … lines in first-seen order."""
    by_num: dict[int, str] = {}
    for line in text.splitlines():
        s = line.strip()
        while s and s[0] in "-*•":
            s = s[1:].lstrip()
        m = re.match(r"^(\d{1,2})\s*[\.)]\s*(.+)$", s)
        if not m:
            continue
        n = int(m.group(1))
        title = m.group(2).strip()
        if 1 <= n <= 15 and len(title) > 2 and n not in by_num:
            by_num[n] = title
    return [by_num[i] for i in range(1, 16) if i in by_num]


def _ordered_lifecycle_steps_from_lines(text: str) -> list[str]:
    """Walk the text in document order; pick canonical ML lifecycle labels."""
    step_patterns: list[tuple[re.Pattern[str], str]] = [
        (re.compile(r"gathering\s+data|data\s+gathering", re.I), "Gathering Data"),
        (re.compile(r"prepar(?:e|ing)(?:\s+that|\s+the)?\s+data|data\s+prepar", re.I), "Preparing that data"),
        (re.compile(r"choosing\s+(?:a\s+)?model|model\s+selection", re.I), "Choosing a model"),
        (re.compile(r"\btraining\b", re.I), "Training"),
        (re.compile(r"\bevaluation\b", re.I), "Evaluation"),
        (re.compile(r"hyper\s*-?\s*parameter(?:\s+tuning)?|hyperparameter\s+tuning", re.I), "Hyper parameter Tuning"),
        (re.compile(r"\bprediction\b", re.I), "Prediction"),
    ]
    seen: set[str] = set()
    ordered: list[str] = []
    for line in text.splitlines():
        for pat, label in step_patterns:
            if label in seen:
                continue
            if pat.search(line):
                ordered.append(label)
                seen.add(label)
                break
    return ordered


def _answer_ml_lifecycle_from_text(body: str, question_lower: str, filename: str) -> Optional[str]:
    """If the question is about ML lifecycle / steps, return a full numbered list from the PDF text."""
    triggers = (
        "life cycle",
        "lifecycle",
        "life-cycle",
        "seven major",
        "seven step",
        "seven stages",
        "7 step",
        "stages of",
        "machine learning project",
        "workflow",
        "ml pipeline",
    )
    if not any(t in question_lower for t in triggers):
        if "step" not in question_lower and "stage" not in question_lower:
            return None

    lower_body = body.lower()
    pos = -1
    for anchor in (
        "machine learning life cycle",
        "machine learning life-cycle",
        "life cycle",
        "lifecycle",
        "ml life cycle",
        "seven major",
        "seven step",
        "machine learning project",
    ):
        i = lower_body.find(anchor)
        if i != -1 and (pos == -1 or i < pos):
            pos = i

    window = body[pos : pos + 12000] if pos != -1 else body[:16000]

    numbered = _numbered_steps_from_text(window)
    if len(numbered) >= 5:
        lines = "\n".join(f"{i + 1}. {title}" for i, title in enumerate(numbered))
        return (
            "Machine learning life cycle involves seven major steps, which are given below:\n"
            f"{lines}\n\n"
            f"(From '{filename}' — numbered items as extracted from the document.)"
        )

    ordered = _ordered_lifecycle_steps_from_lines(window)
    if len(ordered) >= 5:
        lines = "\n".join(f"{i + 1}. {title}" for i, title in enumerate(ordered))
        return (
            "Machine learning life cycle involves seven major steps, which are given below:\n"
            f"{lines}\n\n"
            f"(From '{filename}' — steps detected in document order.)"
        )

    # Last try: scan full body (some PDFs mention steps before the phrase "life cycle")
    numbered_full = _numbered_steps_from_text(body)
    if len(numbered_full) >= 5:
        lines = "\n".join(f"{i + 1}. {title}" for i, title in enumerate(numbered_full))
        return (
            "Machine learning life cycle involves seven major steps, which are given below:\n"
            f"{lines}\n\n"
            f"(From '{filename}' — numbered list from document text.)"
        )

    ordered_full = _ordered_lifecycle_steps_from_lines(body)
    if len(ordered_full) >= 5:
        lines = "\n".join(f"{i + 1}. {title}" for i, title in enumerate(ordered_full))
        return (
            "Machine learning life cycle involves seven major steps, which are given below:\n"
            f"{lines}\n\n"
            f"(From '{filename}' — steps detected in document order.)"
        )

    return None


def local_document_qa_fallback(content, question, filename):
    """Simple extractive fallback when Gemini is unavailable."""
    if not content or not content.strip():
        return (
            "I could not find readable text in this document. "
            "Please upload a clearer PDF or a text-based file."
        )

    normalized_question = question.lower().strip()

    body_text = _extract_primary_document_text(content)
    lifecycle_answer = _answer_ml_lifecycle_from_text(body_text, normalized_question, filename)
    if lifecycle_answer:
        return lifecycle_answer

    # Handle common intent requests (including misspellings like "summery").
    summary_intents = {"summary", "summery", "summarize", "overview", "brief", "gist"}
    explain_intents = {"explain", "points", "key", "important", "main", "topics"}

    def clean_line(text):
        cleaned = text.strip()
        # Remove leading punctuation
        while cleaned and cleaned[0] in {"-", "*", ".", "•", ""}:
            cleaned = cleaned[1:].strip()
        # Remove trailing punctuation
        while cleaned and cleaned[-1] in {".", ":", ",", ";", "!", "?"}:
            cleaned = cleaned[:-1].strip()
        return " ".join(cleaned.split())

    # Basic keyword scoring over lines/paragraphs.
    raw_chunks = []
    for block in content.split("\n\n"):
        for line in block.splitlines():
            cleaned = clean_line(line)
            lower_line = cleaned.lower()
            # Skip obvious OCR noise/header fragments that hurt answer quality.
            if (
                cleaned
                and len(cleaned) > 15
                and "page" not in lower_line
                and "department of" not in lower_line
                and not lower_line.startswith("extracted text:")
                and not lower_line.startswith("file information:")
            ):
                raw_chunks.append(cleaned)

    if not raw_chunks:
        return (
            "I found the document, but there is no readable text to answer from."
        )

    def build_long_explanation(topic, source_lines):
        # Clean the topic - remove trailing punctuation and whitespace
        topic_clean = topic.strip()
        while topic_clean and topic_clean[-1] in {'.', ':', ',', ';', '!', '?'}:
            topic_clean = topic_clean[:-1].strip()
        
        # Use the actual matched lines as the explanation body
        intro = source_lines[0] if source_lines else topic_clean
        # Clean intro as well
        intro_clean = intro.strip()
        while intro_clean and intro_clean[-1] in {'.', ':', ',', ';', '!', '?'}:
            intro_clean = intro_clean[:-1].strip()
        
        # If intro is too short (less than 3 words), try to find a better intro
        if len(intro_clean.split()) < 3 and len(source_lines) > 1:
            for line in source_lines[1:]:
                line_clean = line.strip()
                if len(line_clean.split()) >= 3:
                    intro_clean = line_clean
                    break
        
        support = source_lines[1:5] if len(source_lines) > 1 else []
        # Filter out empty or very short support lines
        support = [line.strip() for line in support if len(line.strip().split()) > 1]
        
        # If we have no support lines but have a decent intro, use intro as a bullet
        if not support and len(intro_clean.split()) > 1:
            support = [intro_clean]
        
        bullets = "\n".join(f"- {line}" for line in support) if support else f"- {intro_clean}"
        
        # If topic is still just "Machine Learning" or similar, make it more descriptive
        if len(topic_clean.split()) <= 2 and support:
            # Use first support line as better topic
            topic_clean = support[0]
            # Truncate if too long
            if len(topic_clean) > 100:
                topic_clean = topic_clean[:97] + "..."
        
        return (
            f"### {topic_clean}\n\n"
            f"**Explanation from the document:**\n"
            f"{intro_clean}\n\n"
            f"**Additional details from the document:**\n{bullets}\n\n"
            f"Ask another concept from the same document and I will explain it similarly."
        )

    if any(token in normalized_question for token in summary_intents):
        summary_points = raw_chunks[:8]
        bullet_points = "\n".join(f"- {line}" for line in summary_points)
        return (
            f"Here are the key points from '{filename}':\n\n"
            f"{bullet_points}\n\n"
            "If you want more detail on any point, ask me to explain that point."
        )

    if "explain" in normalized_question and any(token in normalized_question for token in explain_intents):
        point_index = None
        for token in normalized_question.split():
            if token.isdigit():
                point_index = int(token) - 1
                break

        if point_index is not None and 0 <= point_index < len(raw_chunks):
            topic = raw_chunks[point_index]
            neighbors = raw_chunks[point_index:point_index + 4]
            return (
                f"Here is an explanation for point {point_index + 1} from '{filename}':\n\n"
                f"{build_long_explanation(topic, neighbors)}\n\n"
                "If you want, I can explain another point in the same style."
            )

        explain_points = raw_chunks[:6]
        bullet_points = "\n".join(f"- {line}" for line in explain_points)
        return (
            f"These are the main points from '{filename}':\n\n"
            f"{bullet_points}\n\n"
            "Ask `explain point 1` (or 2, 3...) for more detail on a specific point."
        )

    question_terms = [
        term.lower() for term in question.replace("?", " ").replace(",", " ").split()
        if len(term) > 2 and term.lower() not in {
            "what", "when", "where", "which", "who", "whom", "whose",
            "why", "how", "the", "and", "for", "with", "from", "that",
            "this", "about", "into", "your", "their", "there", "have",
            "has", "had", "are", "was", "were", "can", "could", "would",
            "should", "please", "explain", "tell"
        }
    ]

    scored = []
    for chunk in raw_chunks:
        lower_chunk = chunk.lower()
        score = sum(1 for term in question_terms if term in lower_chunk)
        if score > 0:
            scored.append((score, len(chunk), chunk))

    if not scored:
        preview = " ".join(raw_chunks[:3])[:700]
        return (
            f"I could not find an exact match for your question in '{filename}'. "
            f"Here is a relevant preview from the document:\n\n{preview}"
        )

    # Sort by score (higher first), then by length (longer first for more informative chunks)
    scored.sort(key=lambda item: (-item[0], -item[1]))
    top_chunks = [item[2] for item in scored[:3]]

    if normalized_question.startswith("what is") or normalized_question.startswith("define"):
        topic = top_chunks[0]
        return (
            f"Based on '{filename}', here is a detailed explanation:\n\n"
            f"{build_long_explanation(topic, top_chunks)}"
        )

    combined = "\n".join(f"- {c}" for c in top_chunks)

    return (
        f"Based on '{filename}', these are the most relevant points for your question:\n\n"
        f"{combined}\n\n"
        "For a long chatbot-style response, ask: `summary of this PDF` or `explain point 1`."
    )

@app.route('/api/qa/history/<user_id>', methods=['GET'])
def get_qa_history(user_id):
    """Get user's Q&A history"""
    try:
        user_dir = get_user_data_path(user_id)
        qa_history_file = os.path.join(user_dir, 'qa_history.json')
        
        if os.path.exists(qa_history_file):
            with open(qa_history_file, 'r', encoding='utf-8') as f:
                qa_history = json.load(f)
            
            # Return last 20 sessions
            recent_sessions = qa_history.get("qa_sessions", [])[-20:]
            return jsonify({
                "user_id": user_id,
                "total_sessions": len(qa_history.get("qa_sessions", [])),
                "recent_sessions": recent_sessions
            })
        else:
            return jsonify({
                "user_id": user_id,
                "total_sessions": 0,
                "recent_sessions": []
            })
    
    except Exception as e:
        return jsonify({"error": f"Failed to get Q&A history: {str(e)}"}), 500

@app.route('/api/rag/stats', methods=['GET'])
def get_rag_stats():
    """Get RAG system statistics"""
    try:
        if not rag_system:
            return jsonify({
                "documents_in_rag": 0,
                "total_chunks": 0,
                "vector_dimension": 384,
                "embedding_model": "sentence-transformers/all-MiniLM-L6-v2",
                "retrieval_method": "FAISS Vector Search",
                "chunk_size": 300,
                "chunk_overlap": 75,
                "similarity_threshold": 0.6,
                "rag_enabled": False,
                "system_status": "disabled"
            })

        # Get basic RAG system stats
        stats = {
            "documents_in_rag": rag_system.get_document_count(),
            "total_chunks": rag_system.get_total_chunks(),
            "vector_dimension": rag_system.embedding_dimension,
            "embedding_model": "sentence-transformers/all-MiniLM-L6-v2",
            "retrieval_method": "FAISS Vector Search",
            "chunk_size": 300,
            "chunk_overlap": 75,
            "similarity_threshold": 0.6,
            "rag_enabled": True,
            "system_status": "operational"
        }
        
        return jsonify(stats)
    
    except Exception as e:
        return jsonify({
            "error": f"Failed to get RAG stats: {str(e)}",
            "rag_enabled": False,
            "system_status": "error"
        }), 500

@app.route('/api/real-time-stats', methods=['GET'])
def get_real_time_stats():
    """Get real-time dashboard statistics"""
    try:
        # Count total users
        total_users = len([d for d in os.listdir(USERS_DATA_FOLDER) if os.path.isdir(os.path.join(USERS_DATA_FOLDER, d))])
        
        # Count total documents
        total_documents = 0
        total_questions = 0
        rag_documents = 0
        
        for user_dir in os.listdir(USERS_DATA_FOLDER):
            user_path = os.path.join(USERS_DATA_FOLDER, user_dir)
            if os.path.isdir(user_path):
                # Count user documents
                library_file = os.path.join(user_path, 'library.json')
                if os.path.exists(library_file):
                    with open(library_file, 'r', encoding='utf-8') as f:
                        library = json.load(f)
                        user_docs = len(library.get('documents', []))
                        total_documents += user_docs
                        
                        # Count RAG-enabled documents
                        for doc in library.get('documents', []):
                            if doc.get('rag_processed', False):
                                rag_documents += 1
                
                # Count Q&A sessions
                qa_file = os.path.join(user_path, 'qa_history.json')
                if os.path.exists(qa_file):
                    with open(qa_file, 'r', encoding='utf-8') as f:
                        qa_history = json.load(f)
                        total_questions += len(qa_history.get('qa_sessions', []))
        
        # Calculate processing rates
        rag_adoption_rate = (rag_documents / total_documents * 100) if total_documents > 0 else 0
        
        return jsonify({
            "total_users": total_users,
            "total_documents": total_documents,
            "total_questions": total_questions,
            "rag_documents": rag_documents,
            "rag_adoption_rate": round(rag_adoption_rate, 1),
            "documents_per_user": round(total_documents / total_users, 1) if total_users > 0 else 0,
            "questions_per_document": round(total_questions / total_documents, 1) if total_documents > 0 else 0,
            "timestamp": datetime.now().isoformat(),
            "system_health": "excellent" if rag_adoption_rate > 80 else "good" if rag_adoption_rate > 50 else "fair"
        })
    
    except Exception as e:
        return jsonify({"error": f"Failed to get real-time stats: {str(e)}"}), 500

if __name__ == '__main__':
    print("Starting RAG-Enhanced Document Analysis Backend API...")
    print(f"Upload folder: {UPLOAD_FOLDER}")
    print(f"Users data folder: {USERS_DATA_FOLDER}")
    if model is not None:
        print("✅ Gemini API key loaded (GEMINI_API_KEY from environment)")
    else:
        print("⚠️  Gemini disabled: set GEMINI_API_KEY in python-backend/.env — uploads still work; Q&A uses local text fallback")
    if rag_system is not None:
        print("✅ RAG vector index enabled")
    else:
        print("ℹ️  RAG vector index disabled (rag_system not initialized)")
    print("✅ Document pipeline ready (upload, text extraction, structured local Q&A)")
    print("\nRAG System Features:")
    print("- Vector-based document search")
    print("- Semantic chunk retrieval")
    print("- Retrieval-Augmented Generation")
    print("- Multi-format document support")
    print("- Real-time Q&A with context awareness")
    print("\nAPI Server starting on http://0.0.0.0:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)